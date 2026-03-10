"""
generate_excel.py

Generates App_Roles_Permissions.xlsx with 5 sheets:
  1. Google Workspace
  2. Okta
  3. Azure AD
  4. MS Intune V2
  5. 1Password

Each sheet has columns:
  Role | Permission | Role → Permission Mapping | Is Privileged (T/F) | Description
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = [
    "Role",
    "Permission",
    "Role → Permission Mapping",
    "Is Privileged (T/F)",
    "Description",
]

# ---------------------------------------------------------------------------
# Data definitions
# ---------------------------------------------------------------------------

GOOGLE_WORKSPACE_DATA = [
    {
        "role": "Super Admin",
        "privileged": "TRUE",
        "description": (
            "Has access to all features in the Admin console and Admin API. "
            "Can manage all admins, assign/modify any role. Most privileged role."
        ),
        "permissions": [
            "All Admin Console Settings",
            "User Management (Create/Read/Update/Delete)",
            "Group Management (Create/Read/Update/Delete)",
            "Organizational Unit Management",
            "Domain Management",
            "Billing Management",
            "Security Settings",
            "App Settings (Gmail/Drive/Calendar/etc.)",
            "Reports & Audit Logs",
            "Data Migration",
            "SAML/SSO Configuration",
            "Marketplace Apps Installation",
            "Role Assignment",
            "Service Settings",
            "Device Management",
            "Data Loss Prevention",
        ],
    },
    {
        "role": "Groups Admin",
        "privileged": "FALSE",
        "description": (
            "Can create, manage, and delete Google Groups. "
            "Can manage group members and settings."
        ),
        "permissions": [
            "Groups (Create/Read/Update/Delete)",
            "Group Members (Add/Remove)",
            "Group Settings (Read/Update)",
            "User Profiles (Read)",
            "Organizational Structure (Read)",
        ],
    },
    {
        "role": "User Management Admin",
        "privileged": "FALSE",
        "description": (
            "Can perform most user management tasks for non-admin users. "
            "Cannot manage admins or billing."
        ),
        "permissions": [
            "Users (Create/Read/Update/Delete - non-admin only)",
            "Password Reset (non-admin)",
            "Group Memberships (Read/Update)",
            "Storage Limits (Set)",
            "User Reports (Read)",
            "Shared Drive Reports (Read)",
        ],
    },
    {
        "role": "Help Desk Admin",
        "privileged": "FALSE",
        "description": (
            "Limited to user support tasks like resetting passwords for non-admin users."
        ),
        "permissions": [
            "Password Reset (non-admin users)",
            "User Profiles (Read)",
            "User Activity (Read)",
        ],
    },
    {
        "role": "Services Admin",
        "privileged": "FALSE",
        "description": (
            "Can enable, disable, and configure settings for Google Workspace services."
        ),
        "permissions": [
            "Service Settings (Gmail/Drive/Calendar/Meet/Chat/etc. - Read/Update)",
            "Enable/Disable Services",
            "Application Configuration",
        ],
    },
    {
        "role": "Mobile Admin",
        "privileged": "FALSE",
        "description": (
            "Can manage mobile devices including provisioning, blocking, and wiping."
        ),
        "permissions": [
            "Mobile Devices (Provision/Block/Wipe)",
            "Device Policies (Read/Update)",
            "Device Inventory (Read)",
        ],
    },
    {
        "role": "Storage Admin",
        "privileged": "FALSE",
        "description": (
            "Can monitor and manage Google Drive storage and settings."
        ),
        "permissions": [
            "Storage Monitoring",
            "Storage Management",
            "Data Deletion",
            "Additional Storage Purchase",
            "Drive Storage Settings (Read/Update)",
        ],
    },
    {
        "role": "Google Voice Admin",
        "privileged": "FALSE",
        "description": (
            "Can manage Google Voice setup including phone numbers, licenses, and call settings."
        ),
        "permissions": [
            "Phone Numbers (Manage)",
            "Voice Licenses (Manage)",
            "Call Settings (Read/Update)",
        ],
    },
    {
        "role": "Reseller Admin",
        "privileged": "FALSE",
        "description": (
            "For authorized resellers to manage customer accounts and billing."
        ),
        "permissions": [
            "Customer Account Management",
            "Reseller Billing",
            "Customer Support",
        ],
    },
]

OKTA_DATA = [
    {
        "role": "Super Admin",
        "privileged": "TRUE",
        "description": (
            "Full control over all Okta settings and resources. No restrictions. Most privileged role."
        ),
        "permissions": [
            "All Settings (Full Control)",
            "User Management (Create/Read/Update/Delete/Unlock)",
            "Application Management (All)",
            "Group Management (All)",
            "Policy Management (All)",
            "Security Configuration (All)",
            "API Token Management",
            "Admin Role Assignment",
            "Directory Integrations",
            "Customization (Branding)",
            "Report Generation",
            "System Log Access",
        ],
    },
    {
        "role": "Org Admin",
        "privileged": "TRUE",
        "description": (
            "Can manage most org-wide settings except Super Admin-only settings. "
            "Cannot assign Super Admin role."
        ),
        "permissions": [
            "User Management (Create/Read/Update/Delete/Unlock)",
            "Application Management (All)",
            "Group Management (All)",
            "Policy Management (Most)",
            "Admin Role Assignment (Limited - cannot assign Super Admin)",
            "Directory Integrations",
            "Customization",
            "Report Generation",
            "System Log Access",
        ],
    },
    {
        "role": "Application Admin",
        "privileged": "FALSE",
        "description": (
            "Can manage only assigned applications. No access to org-wide user, "
            "group, or security settings."
        ),
        "permissions": [
            "Assigned Applications (Read/Update/Assign Users/Assign Groups)",
            "Application Settings (Read/Update - assigned only)",
            "Application User Assignment (Add/Remove)",
        ],
    },
    {
        "role": "Group Membership Admin",
        "privileged": "FALSE",
        "description": (
            "Can manage memberships of assigned groups only. "
            "Cannot create, delete groups, or manage group settings."
        ),
        "permissions": [
            "Group Members (Add/Remove - assigned groups only)",
            "Group Membership (Read - assigned groups)",
        ],
    },
    {
        "role": "Help Desk Admin",
        "privileged": "FALSE",
        "description": (
            "User support focus — can unlock accounts, reset passwords, and manage MFA factors."
        ),
        "permissions": [
            "User Account Unlock",
            "Password Reset",
            "MFA Factor (Assign/Revoke)",
            "User Profiles (Read)",
        ],
    },
    {
        "role": "Read Only Admin",
        "privileged": "FALSE",
        "description": (
            "View-only access to the Okta Admin Console. Cannot make any changes."
        ),
        "permissions": [
            "Users (Read)",
            "Groups (Read)",
            "Applications (Read)",
            "Policies (Read)",
            "Reports (Read)",
            "System Log (Read)",
        ],
    },
    {
        "role": "Report Admin",
        "privileged": "FALSE",
        "description": "Can generate and view reports.",
        "permissions": [
            "Reports (Generate/Read)",
            "System Log (Read)",
        ],
    },
    {
        "role": "Group Admin",
        "privileged": "FALSE",
        "description": (
            "Can manage assigned groups including settings and members."
        ),
        "permissions": [
            "Assigned Groups (Read/Update/Delete)",
            "Group Members (Add/Remove - assigned groups)",
            "Group Settings (Read/Update - assigned groups)",
        ],
    },
    {
        "role": "Mobile Admin",
        "privileged": "FALSE",
        "description": (
            "Can manage mobile device-related settings and policies."
        ),
        "permissions": [
            "Mobile Device Management (Read/Update)",
            "Mobile Policies (Read/Update)",
        ],
    },
]

AZURE_AD_DATA = [
    {
        "role": "Global Administrator",
        "privileged": "TRUE",
        "description": (
            "Full access to all administrative features in Microsoft Entra ID and all "
            "Microsoft services. Most privileged role."
        ),
        "permissions": [
            "All Entra ID Settings",
            "User Management (All)",
            "Group Management (All)",
            "Application Registration (All)",
            "Enterprise Application Management (All)",
            "Domain Management",
            "License Management",
            "Security Settings (All)",
            "Conditional Access (All)",
            "Directory Role Assignment",
            "Billing Management",
            "Support Tickets (All)",
            "Service Health (Read)",
        ],
    },
    {
        "role": "User Administrator",
        "privileged": "TRUE",
        "description": (
            "Can create and manage all aspects of users and groups, including resetting "
            "passwords for limited admins."
        ),
        "permissions": [
            "Users (Create/Read/Update/Delete)",
            "Groups (Create/Read/Update/Delete)",
            "Password Reset (limited admins)",
            "User Licenses (Assign/Remove)",
            "Support Tickets (Create/Manage)",
        ],
    },
    {
        "role": "Billing Administrator",
        "privileged": "FALSE",
        "description": "Manages subscriptions and billing-related tasks.",
        "permissions": [
            "Subscriptions (Purchase/Manage)",
            "Billing (Read/Update)",
            "Support Tickets (Billing)",
            "Service Health (Read)",
            "Usage Reports (Read)",
        ],
    },
    {
        "role": "Security Administrator",
        "privileged": "TRUE",
        "description": (
            "Has read and management capabilities for security features and configurations."
        ),
        "permissions": [
            "Security Features (Read/Manage)",
            "Conditional Access Policies (Read/Update)",
            "Security Alerts (Read/Manage)",
            "Identity Protection (Read/Update)",
            "Privileged Identity Management (Read)",
            "Security Reports (Read)",
        ],
    },
    {
        "role": "Helpdesk Administrator",
        "privileged": "FALSE",
        "description": (
            "Can reset passwords for non-administrators and Helpdesk Administrators."
        ),
        "permissions": [
            "Password Reset (non-admin users)",
            "User Profiles (Read)",
            "Support Tickets (Create/Manage)",
            "Service Health (Read)",
        ],
    },
    {
        "role": "Application Administrator",
        "privileged": "FALSE",
        "description": (
            "Can create and manage all aspects of app registrations and enterprise apps."
        ),
        "permissions": [
            "App Registrations (Create/Read/Update/Delete)",
            "Enterprise Apps (Read/Update)",
            "Application Proxy (Configure)",
            "Single Sign-On (Configure)",
            "User/Group Assignment to Apps",
        ],
    },
    {
        "role": "Cloud Application Administrator",
        "privileged": "FALSE",
        "description": (
            "Same as Application Administrator except cannot manage Application Proxy."
        ),
        "permissions": [
            "App Registrations (Create/Read/Update/Delete)",
            "Enterprise Apps (Read/Update)",
            "Single Sign-On (Configure)",
            "User/Group Assignment to Apps",
        ],
    },
    {
        "role": "Conditional Access Administrator",
        "privileged": "TRUE",
        "description": (
            "Can manage Conditional Access settings and policies."
        ),
        "permissions": [
            "Conditional Access Policies (Create/Read/Update/Delete)",
            "Named Locations (Manage)",
            "Sign-in Logs (Read)",
        ],
    },
    {
        "role": "Exchange Administrator",
        "privileged": "FALSE",
        "description": "Can manage all aspects of Exchange Online.",
        "permissions": [
            "Exchange Online Settings (All)",
            "Mailboxes (Manage)",
            "Mail Flow Rules (Manage)",
            "Distribution Groups (Manage)",
            "Quarantine (Manage)",
        ],
    },
    {
        "role": "SharePoint Administrator",
        "privileged": "FALSE",
        "description": "Can manage all aspects of SharePoint Online.",
        "permissions": [
            "SharePoint Online Settings (All)",
            "Site Collections (Manage)",
            "OneDrive (Manage)",
            "Sharing Policies (Manage)",
        ],
    },
    {
        "role": "Teams Administrator",
        "privileged": "FALSE",
        "description": "Can manage all aspects of Microsoft Teams.",
        "permissions": [
            "Teams Settings (All)",
            "Teams Policies (Manage)",
            "Meeting Policies (Manage)",
            "Messaging Policies (Manage)",
            "Voice Settings (Manage)",
        ],
    },
    {
        "role": "Compliance Administrator",
        "privileged": "FALSE",
        "description": (
            "Can read and manage compliance configuration and reports."
        ),
        "permissions": [
            "Compliance Settings (Read/Manage)",
            "Compliance Reports (Read)",
            "Data Loss Prevention (Manage)",
            "Retention Policies (Manage)",
            "Audit Logs (Read)",
        ],
    },
    {
        "role": "Security Reader",
        "privileged": "FALSE",
        "description": "Read-only access to security features.",
        "permissions": [
            "Security Features (Read)",
            "Conditional Access (Read)",
            "Security Alerts (Read)",
            "Identity Protection (Read)",
            "Security Reports (Read)",
            "Audit Logs (Read)",
        ],
    },
    {
        "role": "Global Reader",
        "privileged": "FALSE",
        "description": (
            "Read-only version of Global Administrator. Can read all settings but cannot make changes."
        ),
        "permissions": [
            "All Entra ID Settings (Read Only)",
            "Users (Read)",
            "Groups (Read)",
            "Applications (Read)",
            "Policies (Read)",
            "Reports (Read)",
        ],
    },
    {
        "role": "Privileged Role Administrator",
        "privileged": "TRUE",
        "description": (
            "Can manage role assignments in Microsoft Entra ID and Privileged Identity Management."
        ),
        "permissions": [
            "Role Assignments (Create/Read/Update/Delete)",
            "PIM Settings (Manage)",
            "Consent Requests (Manage)",
            "Directory Roles (Manage)",
        ],
    },
]

INTUNE_DATA = [
    {
        "role": "Intune Administrator",
        "privileged": "TRUE",
        "description": (
            "Full access to all Intune features and settings. Most privileged Intune role."
        ),
        "permissions": [
            "All Intune Resources (Create/Read/Update/Delete/Assign)",
            "Device Management (All)",
            "App Management (All)",
            "Policy Management (All)",
            "Reports (All)",
            "Enrollment (All)",
            "Remote Actions (All)",
            "Security Baselines (All)",
            "Endpoint Security (All)",
            "Audit Logs (Read)",
        ],
    },
    {
        "role": "Application Manager",
        "privileged": "FALSE",
        "description": "Can manage mobile and managed applications.",
        "permissions": [
            "Managed Apps (Assign/Create/Delete/Read/Update/Wipe)",
            "Mobile Apps (Assign/Create/Delete/Read/Relate/Update)",
            "Device Configurations (Read)",
            "Device Information (Read)",
            "Deployment Plans (Assign/Create/Delete/Read/Update)",
            "Organization (Read)",
        ],
    },
    {
        "role": "Endpoint Security Manager",
        "privileged": "TRUE",
        "description": "Can manage security and compliance features.",
        "permissions": [
            "Compliance Policies (Assign/Create/Delete/Read/Update)",
            "Device Configurations (Read)",
            "Security Baselines (Assign/Create/Delete/Read/Update)",
            "Microsoft Defender ATP (Read)",
            "Managed Devices (Read)",
            "Endpoint Security Policies (Assign/Create/Delete/Read/Update)",
            "Audit Data (Read)",
        ],
    },
    {
        "role": "Help Desk Operator",
        "privileged": "FALSE",
        "description": (
            "Provides support to end users with remote tasks and troubleshooting."
        ),
        "permissions": [
            "Remote Tasks (Execute - Remote Lock/Reset Passcode/Retire)",
            "Device Inventory (Read)",
            "Devices (Read)",
            "Users (Read)",
            "Mobile Apps (Assign)",
            "Compliance Policies (Read)",
            "Device Configurations (Read)",
            "Enrollment Programs (Read)",
        ],
    },
    {
        "role": "School Administrator",
        "privileged": "FALSE",
        "description": (
            "Manages school/classroom Windows devices in Intune for Education."
        ),
        "permissions": [
            "Student/Teacher Devices (Assign/Read)",
            "Device Reset/Retire",
            "Classroom Device Configuration",
            "Enrollment (School devices)",
        ],
    },
    {
        "role": "Policy and Profile Manager",
        "privileged": "FALSE",
        "description": (
            "Can manage compliance policies, configuration profiles, and enrollment."
        ),
        "permissions": [
            "Compliance Policies (Assign/Create/Delete/Read/Update)",
            "Device Configurations (Assign/Create/Delete/Read/Update)",
            "Enrollment Programs (Assign/Create/Delete/Read/Update)",
            "Organization (Read)",
            "Audit Data (Read)",
        ],
    },
    {
        "role": "Read Only Operator",
        "privileged": "FALSE",
        "description": "View-only access to all Intune resources.",
        "permissions": [
            "All Intune Resources (Read Only)",
            "Audit Data (Read)",
        ],
    },
    {
        "role": "Intune Role Administrator",
        "privileged": "TRUE",
        "description": "Can manage Intune RBAC roles and assignments.",
        "permissions": [
            "Intune Roles (Assign/Create/Delete/Read/Update)",
            "Role Assignments (Create/Delete/Read/Update)",
        ],
    },
]

ONEPASSWORD_DATA = [
    {
        "role": "Owner",
        "privileged": "TRUE",
        "description": (
            "Has all administrative rights including billing and account deletion. "
            "There must always be at least one Owner."
        ),
        "permissions": [
            "Create Vaults",
            "View Administrative Sidebar",
            "Recover Accounts",
            "Manage People",
            "Suspend People",
            "Invite & Remove People",
            "Manage All Groups",
            "Manage Settings",
            "Manage Billing",
            "Delete Account",
        ],
    },
    {
        "role": "Administrator",
        "privileged": "TRUE",
        "description": (
            "Can manage vaults, groups, members, and account recovery. "
            "Cannot manage billing or delete the account."
        ),
        "permissions": [
            "Create Vaults",
            "View Administrative Sidebar",
            "Recover Accounts",
            "Manage People",
            "Suspend People",
            "Invite & Remove People",
            "Manage All Groups",
            "Manage Settings",
        ],
    },
    {
        "role": "Security",
        "privileged": "FALSE",
        "description": (
            "Can view security reports for auditing purposes. "
            "Typically used for security teams or auditors."
        ),
        "permissions": [
            "View Administrative Sidebar",
            "View Security Reports",
        ],
    },
    {
        "role": "Team Member",
        "privileged": "FALSE",
        "description": (
            "Default group for all users except guests. Can access assigned vaults "
            "but has no administrative capabilities."
        ),
        "permissions": [
            "Create Vaults (teams created after April 9 2017)",
            "Access Assigned Vaults",
        ],
    },
    {
        "role": "Provision Managers",
        "privileged": "FALSE",
        "description": (
            "Used for automated provisioning via SCIM/CLI. "
            "Can provision and confirm new team members."
        ),
        "permissions": [
            "Provision New Members",
            "Confirm New Members",
            "Access Employee Vaults (before user signup)",
        ],
    },
    {
        "role": "Recovery Group",
        "privileged": "FALSE",
        "description": (
            "Designated account recovery managers who can restore access for users "
            "who lose credentials."
        ),
        "permissions": [
            "Recover Accounts",
            "Account Recovery Management",
        ],
    },
    {
        "role": "Guest",
        "privileged": "FALSE",
        "description": (
            "External users with limited access to specific shared vaults only."
        ),
        "permissions": [
            "Access Specific Shared Vaults (Read Only by default)",
        ],
    },
]

SHEETS = [
    ("Google Workspace", GOOGLE_WORKSPACE_DATA),
    ("Okta", OKTA_DATA),
    ("Azure AD", AZURE_AD_DATA),
    ("MS Intune V2", INTUNE_DATA),
    ("1Password", ONEPASSWORD_DATA),
]

# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# Column widths (approximate, in character units)
COL_WIDTHS = [30, 55, 26, 22, 80]


def build_rows(role_data):
    """Expand role data into flat rows suitable for the worksheet."""
    rows = []
    for entry in role_data:
        role = entry["role"]
        privileged = entry["privileged"]
        description = entry["description"]
        permissions = entry["permissions"]
        for i, perm in enumerate(permissions):
            rows.append(
                [
                    role if i == 0 else "",        # Role (only on first permission row)
                    perm,                           # Permission
                    "Yes",                          # Role → Permission Mapping
                    privileged if i == 0 else "",   # Is Privileged (only on first row)
                    description if i == 0 else "",  # Description (only on first row)
                ]
            )
    return rows


def write_sheet(wb, sheet_name, role_data):
    ws = wb.create_sheet(title=sheet_name)

    # Write headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Write data rows
    rows = build_rows(role_data)
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Set column widths
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Set row height for header
    ws.row_dimensions[1].height = 30


def main():
    wb = openpyxl.Workbook()
    # Remove the default blank sheet
    wb.remove(wb.active)

    for sheet_name, role_data in SHEETS:
        write_sheet(wb, sheet_name, role_data)

    output_path = "App_Roles_Permissions.xlsx"
    wb.save(output_path)
    print(f"Generated: {output_path}")


if __name__ == "__main__":
    main()
